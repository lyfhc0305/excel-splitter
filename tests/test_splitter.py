import contextlib
import io
import sys
import tempfile
import threading
import time
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils.datetime import CALENDAR_MAC_1904

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / 'src'))
import excel_splitter as app
import split_core as core


class FormulaTests(unittest.TestCase):
    def test_formula_tokens_and_absolute_references(self):
        result = core.rebuild_formula('=IF(A4="A4",LOG10($B$4),SUM(B1:B6))', {1: 1, 4: 2, 6: 3}, 1, 6)
        self.assertEqual(result, '=IF(A2="A4",LOG10($B$2),SUM(B1:B3))')

    def test_deleted_refs_are_errors_not_other_groups(self):
        self.assertEqual(core.rebuild_formula('=B3+SUM(B2:B3)', {1: 1, 4: 2}), '=#REF!+SUM(#REF!)')

    def test_whole_axis_ranges(self):
        self.assertEqual(core.rebuild_formula('=SUM($2:$6)+SUM(B:B)', {1: 1, 3: 2, 6: 3}), '=SUM($2:$3)+SUM(B:B)')
        self.assertEqual(core.rebuild_formula_by_columns('=SUM($B:$F)+SUM(2:6)', {1: 1, 3: 2, 6: 3}), '=SUM($B:$C)+SUM(2:6)')

    def test_escaped_unicode_sheet_and_other_sheet(self):
        self.assertEqual(core.rebuild_formula("='财务''部'!B4", {4: 1}, sheet_name="财务'部"), "='财务''部'!B1")
        with self.assertRaisesRegex(ValueError, '其他工作表'):
            core.rebuild_formula('=Other!B4', {4: 1}, sheet_name='Sheet')

    def test_dynamic_reference_is_not_silently_corrupted(self):
        with self.assertRaisesRegex(ValueError, '动态引用'):
            core.rebuild_formula('=INDIRECT("B4")', {4: 1})


class SplitTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.source = self.root / '源表.xlsx'
        self.out = self.root / 'out'

    def tearDown(self):
        self.temp.cleanup()

    def save(self, rows, title='Sheet'):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        for row in rows:
            ws.append(row)
        wb.save(self.source)
        return wb, ws

    def split(self, **kwargs):
        return app.split_workbook(self.source, None, kwargs.pop('header_rows', 1), 1, self.out, **kwargs)

    def test_totals_fixed_rows_styles_and_dates_roundtrip(self):
        wb, ws = self.save([['部门', '值'], ['甲', 10], ['乙', 20], ['甲', 30], ['合计', '=SUM(B2:B4)']])
        wb.epoch = CALENDAR_MAC_1904
        wb.calculation.calcMode = 'manual'
        ws['B4'].font = Font(name='Arial', bold=True, color='FF112233')
        ws['B4'].number_format = '0.00'
        ws.row_dimensions[4].height = 35
        ws.row_dimensions[4].font = Font(italic=True)
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['B'].font = Font(bold=True)
        ws.print_area = 'A1:B5'
        ws.print_title_rows = '1:1'
        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = 'A1:B4'
        ws.protection.sheet = True
        ws.row_breaks.append(Break(id=4))
        wb.save(self.source)
        files = self.split(footer_rows=1)
        self.assertEqual(len(files), 2)
        result = openpyxl.load_workbook(files[0])
        try:
            target = result.active
            self.assertEqual(target['B4'].value, '=SUM(B2:B3)')
            self.assertEqual(target['B3'].value, 30)
            self.assertTrue(target['B3'].font.bold)
            self.assertEqual(target['B3'].number_format, '0.00')
            self.assertEqual(target.row_dimensions[3].height, 35)
            self.assertTrue(target.row_dimensions[3].font.italic)
            self.assertTrue(target.column_dimensions['B'].font.bold)
            self.assertIn('$A$1:$B$4', target.print_area)
            self.assertEqual(target.auto_filter.ref, 'A1:B3')
            self.assertEqual(target.freeze_panes, 'B2')
            self.assertTrue(target.protection.sheet)
            self.assertEqual(target.row_breaks.brk[0].id, 3)
            self.assertEqual(result.epoch, wb.epoch)
            self.assertEqual(result.calculation.calcMode, 'auto')
        finally:
            result.close()

    def test_filename_collisions_and_existing_files(self):
        self.save([['部门'], ['A/B'], ['A:B'], ['abc'], ['ABC'], ['x' * 100 + '1'], ['x' * 100 + '2']])
        first = self.split()
        originals = {path: path.read_bytes() for path in first}
        second = self.split()
        self.assertEqual(len(set(path.name.casefold() for path in first + second)), 12)
        self.assertTrue(all(path.read_bytes() == data for path, data in originals.items()))

    def test_blanks_are_kept_and_distinct_from_literal_label(self):
        self.save([['部门', '值'], [None, 10], ['（空白关键字）', 20], [' ', 30]])
        files = self.split()
        values = []
        for file in files:
            wb = openpyxl.load_workbook(file)
            values.extend(row[1] for row in wb.active.iter_rows(min_row=2, values_only=True))
            wb.close()
        self.assertEqual(sorted(values), [10, 20, 30])
        self.assertEqual(len(files), 2)

    def test_merge_anchor_survives_removed_row(self):
        wb, ws = self.save([['部门', '说明'], ['甲', '共用说明'], ['乙'], ['乙']])
        ws.merge_cells('B2:B4')
        ws['B2'].font = Font(bold=True)
        wb.save(self.source)
        files = self.split()
        wb = openpyxl.load_workbook(files[1])
        self.assertEqual(wb.active['B2'].value, '共用说明')
        self.assertTrue(wb.active['B2'].font.bold)
        self.assertIn('B2:B3', wb.active.merged_cells)
        wb.close()

    def test_column_split_totals_merges_and_filter_ids(self):
        wb, ws = self.save([['项目', '甲', '乙', '甲', '合计'], ['销量', 2, 3, 5, '=SUM(B2:D2)'], ['标题', '=B2+D2']])
        ws.merge_cells('B3:D3')
        ws.auto_filter.ref = 'A1:E2'
        ws.auto_filter.add_filter_column(3, ['5'])
        ws.column_dimensions.group('B', 'D', hidden=True)
        ws.print_area = 'A1:E3'
        wb.save(self.source)
        files = app.split_workbook_by_row(self.source, None, 1, 1, self.out, 1)
        wb = openpyxl.load_workbook(files[0])
        self.assertEqual(wb.active['D2'].value, '=SUM(B2:C2)')
        self.assertEqual(wb.active['B3'].value, '=B2+C2')
        self.assertEqual(wb.active.auto_filter.filterColumn[0].colId, 2)
        self.assertTrue(wb.active.column_dimensions['C'].hidden)
        self.assertIn('$A$1:$D$3', wb.active.print_area)
        wb.close()

    def test_no_header_preview_and_output_agree(self):
        self.save([['甲', 1], ['乙', 2], ['甲', 3]])
        preview = app.build_sheet_preview(self.source, 'Sheet', 0, 1)
        self.assertEqual(preview['split_objects'], [{'name': '甲', 'count': 2}, {'name': '乙', 'count': 1}])
        self.assertEqual(len(self.split(header_rows=0)), 2)

    def test_invalid_parameters_rejected_by_preview_and_split(self):
        self.save([['标题', '值'], ['甲', 1]])
        for header, key, footer in [(-1, 1, 0), (1, 0, 0), (1, 3, 0), (1, 1, -1), (1, 1, 1)]:
            for function in (app.split_workbook, app.build_sheet_preview):
                with self.subTest(header=header, key=key, footer=footer, function=function):
                    with self.assertRaises(ValueError):
                        if function is app.split_workbook:
                            function(self.source, 'Sheet', header, key, self.out, footer)
                        else:
                            function(self.source, 'Sheet', header, key, footer)

    def test_formula_group_key_is_rejected(self):
        self.save([['部门'], ['="甲"']])
        with self.assertRaisesRegex(ValueError, 'A2'):
            self.split()

    def test_hidden_sheet_becomes_visible(self):
        wb, ws = self.save([['部门'], ['甲']])
        wb.create_sheet('Visible')
        ws.sheet_state = 'hidden'
        wb.save(self.source)
        file = self.split()[0]
        wb = openpyxl.load_workbook(file)
        self.assertEqual(wb.active.sheet_state, 'visible')
        wb.close()

    def test_literal_formula_like_text_is_not_executed(self):
        wb, ws = self.save([['部门', '内容'], ['甲', '=hello']])
        ws['B2'].data_type = 's'
        wb.save(self.source)
        file = self.split()[0]
        wb = openpyxl.load_workbook(file)
        self.assertEqual(wb.active['B2'].data_type, 's')
        self.assertEqual(wb.active['B2'].value, '=hello')
        wb.close()

    def test_late_save_failure_publishes_nothing(self):
        self.save([['部门'], ['甲'], ['乙']])
        original = openpyxl.Workbook.save
        count = 0
        def failing_save(wb, path):
            nonlocal count
            count += 1
            if count == 2:
                raise OSError('disk full')
            return original(wb, path)
        with patch.object(openpyxl.Workbook, 'save', failing_save):
            with self.assertRaises(OSError):
                self.split()
        self.assertEqual(list(self.out.iterdir()), [])

    def test_rules_and_named_range_follow_rows(self):
        wb, ws = self.save([['部门', '值'], ['甲', 2], ['乙', 3], ['甲', '=SUM(Amounts)']])
        wb.defined_names.add(DefinedName('Amounts', attr_text="'Sheet'!$B$2:$B$3"))
        wb.defined_names.add(DefinedName('Unrelated', attr_text="'Missing'!$A$1"))
        dv = DataValidation(type='whole', operator='between', formula1='0', formula2='10')
        dv.add('B2:B4')
        ws.add_data_validation(dv)
        ws.conditional_formatting.add('B2:B4', FormulaRule(formula=['B2>1'], fill=PatternFill('solid', fgColor='FFFF0000')))
        wb.save(self.source)
        file = self.split()[0]
        wb = openpyxl.load_workbook(file)
        self.assertEqual(wb.defined_names['Amounts'].attr_text, "'Sheet'!$B$2:$B$2")
        self.assertNotIn('Unrelated', wb.defined_names)
        self.assertEqual(str(wb.active.data_validations.dataValidation[0].sqref), 'B2:B3')
        self.assertEqual(str(next(iter(wb.active.conditional_formatting)).sqref), 'B2:B3')
        wb.close()

    def test_template_outputs_real_xlsx(self):
        wb, ws = self.save([['部门'], ['甲']])
        wb.template = True
        self.source = self.source.with_suffix('.xltx')
        wb.save(self.source)
        wb = openpyxl.load_workbook(self.split()[0])
        self.assertFalse(wb.template)
        wb.close()

    def test_unsupported_features_do_not_leave_outputs(self):
        wb, ws = self.save([['部门', '值'], ['甲', '=Other!A1']])
        with self.assertRaisesRegex(ValueError, '其他工作表'):
            self.split()
        self.assertEqual(list(self.out.iterdir()), [])

    def test_missing_sheet_closes_workbook(self):
        self.save([['部门'], ['甲']])
        original = app.close_workbook_compatible
        with patch.object(app, 'close_workbook_compatible', wraps=original) as close:
            with self.assertRaisesRegex(ValueError, '找不到工作表'):
                app.split_workbook(self.source, '不存在', 1, 1, self.out)
            close.assert_called_once()

    def test_publish_failure_rolls_back_only_this_run(self):
        self.save([['部门'], ['甲'], ['乙']])
        self.out.mkdir()
        existing = self.out / 'keep.xlsx'
        existing.write_bytes(b'original')
        original = Path.open
        count = 0
        def failing_open(path, mode='r', *args, **kwargs):
            nonlocal count
            if mode == 'xb':
                count += 1
                if count == 2:
                    raise PermissionError('locked')
            return original(path, mode, *args, **kwargs)
        with patch.object(Path, 'open', failing_open):
            with self.assertRaises(PermissionError):
                self.split()
        self.assertEqual(list(self.out.iterdir()), [existing])
        self.assertEqual(existing.read_bytes(), b'original')

    def test_missing_formula_reference_is_reported(self):
        self.save([['部门', '值'], ['甲', '=B3'], ['乙', 10]])
        result = self.split()
        self.assertEqual(len(result.warnings), 1)
        self.assertIn('B2', result.warnings[0])

    def test_merged_key_and_zero_fixed_columns(self):
        wb, ws = self.save([['甲', None, '乙'], [1, 2, 3]])
        ws.merge_cells('A1:B1')
        wb.save(self.source)
        preview = app.build_sheet_preview_by_row(self.source, 'Sheet', 0, 1)
        files = app.split_workbook_by_row(self.source, None, 0, 1, self.out)
        self.assertEqual([item['count'] for item in preview['split_objects']], [2, 1])
        wb = openpyxl.load_workbook(files[0])
        self.assertEqual(list(wb.active.values), [('甲', None), (1, 2)])
        wb.close()

    def test_preview_size_is_bounded(self):
        wb, ws = self.save([['甲']])
        ws.cell(5000, 1000, '乙')
        wb.save(self.source)
        preview = app.build_sheet_preview(self.source, 'Sheet', 4000, 1000)
        self.assertLess(len(preview['column_headers']), 60)
        self.assertLess(len(preview['preview_rows']), 60)
        self.assertEqual(sum(item['count'] for item in preview['split_objects']), 1000)

    def test_conversion_failure_cleans_temp_directory(self):
        self.source = self.source.with_suffix('.xls')
        self.source.write_bytes(b'old-format')
        converted = tempfile.TemporaryDirectory(dir=self.root)
        directory = Path(converted.name)
        with patch.object(app, 'convert_to_xlsx', return_value=(directory / 'bad.xlsx', converted)):
            with self.assertRaises(FileNotFoundError):
                app.load_workbook_compatible(self.source)
        self.assertFalse(directory.exists())

    def test_converter_timeout_tries_fallback(self):
        def fallback(source, output):
            result = output / 'converted.xlsx'
            result.write_bytes(b'converted')
            return result
        with patch.object(app, 'convert_with_libreoffice', side_effect=app.subprocess.TimeoutExpired('lo', 120)):
            with patch.object(app, 'convert_with_windows_com', side_effect=fallback) as convert:
                result, temporary = app.convert_to_xlsx(self.source)
                try:
                    self.assertEqual(result.read_bytes(), b'converted')
                    convert.assert_called_once()
                finally:
                    temporary.cleanup()


class InterfaceTests(unittest.TestCase):
    def test_cli_zero_headers_and_missing_parameters(self):
        self.assertEqual(app.parse_args(['--input', 'x.xlsx', '--header-rows', '0', '--key-column', '1']).header_rows, 0)
        with contextlib.redirect_stderr(io.StringIO()):
            with self.assertRaises(SystemExit) as error:
                app.parse_args(['--input', 'x.xlsx'])
        self.assertEqual(error.exception.code, 2)

    def test_windows_filename_rules(self):
        for name in ['CON', 'aux.txt', 'LPT1', 'com¹']:
            self.assertTrue(core.safe_file_name(name).startswith('_'))
        self.assertEqual(core.safe_file_name('x. '), 'x')
        self.assertEqual(core.safe_file_name('a\x01b'), 'a_b')


class BackgroundInterfaceTests(unittest.TestCase):
    def setUp(self):
        try:
            self.root = app.tk.Tk()
        except app.tk.TclError as exc:
            self.skipTest(f'Tk display unavailable: {exc}')
        self.root.withdraw()
        with patch.object(app.tk, 'Tk', return_value=self.root):
            self.gui = app.SplitterApp()

    def tearDown(self):
        if hasattr(self, 'gui'):
            self.gui._splitting = False
            self.gui._close()

    def pump_until(self, condition):
        deadline = time.monotonic() + 5
        while not condition() and time.monotonic() < deadline:
            self.root.update()
            time.sleep(0.005)
        self.assertTrue(condition(), 'background result did not reach the interface')

    def test_background_job_keeps_ui_responsive_and_discards_stale_preview(self):
        started, release = threading.Event(), threading.Event()
        delivered, responsive = [], []
        def slow():
            started.set()
            release.wait(3)
            return 'old'
        try:
            self.gui._start_job('preview', slow, lambda value, error: delivered.append(value))
            self.pump_until(started.is_set)
            self.root.after(1, lambda: responsive.append(True))
            self.pump_until(lambda: bool(responsive))
            self.gui._start_job('preview', lambda: 'new', lambda value, error: delivered.append(value))
            release.set()
            self.pump_until(lambda: bool(delivered))
            self.assertEqual(delivered, ['new'])
        finally:
            release.set()

    def test_failed_split_restores_controls(self):
        self.gui.input_var.set('nonexistent.xlsx')
        with patch.object(app.messagebox, 'showerror') as show:
            self.gui.run_split()
            self.assertTrue(self.gui._splitting)
            self.pump_until(lambda: not self.gui._splitting)
            show.assert_called_once()
            self.assertEqual(str(self.gui.sheet_combo.cget('state')), 'readonly')

    def test_key_row_highlight_is_rendered(self):
        self.gui.render_preview_table(['A列/1: 甲'], [{'row_no': 1, 'kind': 'keyrow', 'values': ['甲']}])
        item = self.gui.preview_table.get_children()[0]
        self.assertEqual(self.gui.preview_table.item(item, 'tags'), ('keyrow',))


if __name__ == '__main__':
    unittest.main()
