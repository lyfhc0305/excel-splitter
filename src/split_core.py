"""Shared, loss-aware row/column splitting engine."""
from __future__ import annotations

import copy
import os
import re
import tempfile
from itertools import chain
from bisect import bisect_left, bisect_right
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.tokenizer import Tokenizer, TokenizerError
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange


CELL = re.compile(r"(\$?)([A-Za-z]{1,3})(\$?)([1-9][0-9]*)\Z")
AXIS = re.compile(r"(\$?)([A-Za-z]{1,3}|[1-9][0-9]*)\Z")


def normalize_group_value(value):
    if value is None:
        return None
    return str(value).strip() or None


def safe_file_name(name):
    cleaned = re.sub(r'[\x00-\x1f\\/:*?"<>|]+', '_', str(name).strip())
    cleaned = cleaned[:80].rstrip(' .') or '未命名'
    if re.fullmatch(r'(CON|PRN|AUX|NUL|COM[1-9¹²³]|LPT[1-9¹²³])(?:\..*)?', cleaned, re.I):
        cleaned = '_' + cleaned
    return cleaned


def validate_parameters(ws, leading, key, trailing, by_columns=False):
    labels = ('左侧固定列数', '关键行序号', '右侧固定列数') if by_columns else ('表头行数', '关键列序号', '表尾行数')
    for value, label in zip((leading, key, trailing), labels):
        if isinstance(value, bool) or not isinstance(value, int):
            raise ValueError(f'{label}必须是整数。')
    if leading < 0 or trailing < 0:
        raise ValueError(f'{labels[0]}和{labels[2]}不能为负数。')
    size = ws.max_column if by_columns else ws.max_row
    key_size = ws.max_row if by_columns else ws.max_column
    if leading + trailing >= size:
        raise ValueError(f'{labels[0]}与{labels[2]}之和必须小于工作表总{"列" if by_columns else "行"}数。')
    if not 1 <= key <= key_size:
        raise ValueError(f'{labels[1]}必须在 1 到 {key_size} 之间。')


def collect_groups(ws, leading, key, trailing=0, by_columns=False):
    validate_parameters(ws, leading, key, trailing, by_columns)
    size = ws.max_column if by_columns else ws.max_row
    merged = {}
    for region in ws.merged_cells.ranges:
        lo, hi = (region.min_row, region.max_row) if by_columns else (region.min_col, region.max_col)
        if lo <= key <= hi:
            start, end = (region.min_col, region.max_col) if by_columns else (region.min_row, region.max_row)
            anchor = ws.cell(region.min_row, region.min_col)
            for index in range(max(start, leading + 1), min(end, size - trailing) + 1):
                merged[index] = anchor
    groups = {}
    for index in range(leading + 1, size - trailing + 1):
        cell = merged.get(index)
        if cell is None:
            cell = ws.cell(key, index) if by_columns else ws.cell(index, key)
        if cell.data_type in ('f', 'e'):
            raise ValueError(f'分组关键单元格 {cell.coordinate} 包含公式或错误值，请先将关键字转换为有效的固定值。')
        value = normalize_group_value(cell.value)
        groups.setdefault(value, []).append(index)
    # A blank key is an explicit group, distinct even from an identical literal label.
    blank_label = '（空白关键字）'
    while blank_label in groups:
        blank_label += '_'
    return {(blank_label if name is None else name): indices for name, indices in groups.items()}


def collect_group_rows(ws, header_rows, key_column, footer_rows=0):
    return collect_groups(ws, header_rows, key_column, footer_rows)


def collect_group_columns(ws, header_cols, key_row, footer_cols=0):
    return collect_groups(ws, header_cols, key_row, footer_cols, True)


class ReferenceMapper:
    def __init__(self, mapping, by_columns=False, sheet_name=None):
        self.mapping = mapping
        self.indices = sorted(mapping)
        self.by_columns = by_columns
        self.sheet_name = sheet_name
        self.names = set()

    def interval(self, start, end):
        lo, hi = sorted((start, end))
        left = bisect_left(self.indices, lo)
        right = bisect_right(self.indices, hi)
        if left == right:
            return None
        result = self.mapping[self.indices[left]], self.mapping[self.indices[right - 1]]
        return result if start <= end else result[::-1]

    def reference(self, text):
        prefix = ''
        if '!' in text:
            sheet, text = text.rsplit('!', 1)
            decoded = sheet[1:-1].replace("''", "'") if sheet.startswith("'") and sheet.endswith("'") else sheet
            if self.sheet_name is None or decoded.casefold() != self.sheet_name.casefold():
                raise ValueError(f'公式引用了其他工作表或外部文件：{sheet}!{text}。请先将这类引用转换为值后拆分。')
            prefix = sheet + '!'
        parts = text.split(':')
        matches = [CELL.fullmatch(part) for part in parts]
        if len(parts) <= 2 and all(matches):
            values = []
            for match in matches:
                col = column_index_from_string(match[2].upper())
                row = int(match[4])
                if col > 16384 or row > 1048576:
                    return prefix + text  # A defined name, not a cell address.
                values.append(col if self.by_columns else row)
            mapped = self.interval(*values) if len(values) == 2 else ((self.mapping[values[0]],) if values[0] in self.mapping else None)
            if mapped is None:
                return '#REF!'
            result = []
            for match, value in zip(matches, mapped):
                col = get_column_letter(value) if self.by_columns else match[2]
                row = match[4] if self.by_columns else str(value)
                result.append(f'{match[1]}{col}{match[3]}{row}')
            return prefix + ':'.join(result)
        # Whole-row and whole-column references also move when their axis is split.
        matches = [AXIS.fullmatch(part) for part in parts]
        if len(parts) == 2 and all(matches):
            is_columns = all(m[2].isalpha() for m in matches)
            is_rows = all(m[2].isdigit() for m in matches)
            if is_columns or is_rows:
                if is_columns != self.by_columns:
                    return prefix + text
                values = [column_index_from_string(m[2].upper()) if is_columns else int(m[2]) for m in matches]
                mapped = self.interval(*values)
                if mapped is None:
                    return '#REF!'
                return prefix + ':'.join(m[1] + (get_column_letter(v) if is_columns else str(v)) for m, v in zip(matches, mapped))
        if '[' in text or ']' in text or prefix:
            raise ValueError(f'暂不支持此公式引用：{prefix}{text}。请先转换为值。')
        self.names.add(text)
        return text  # Named references are mapped separately.

    def formula(self, formula):
        if not isinstance(formula, str):
            raise ValueError('暂不支持数组公式或数据表公式，请先转换为值。')
        had_equals = formula.startswith('=')
        try:
            tokens = Tokenizer(formula if had_equals else '=' + formula).items
        except TokenizerError as exc:
            raise ValueError(f'无法安全解析公式：{formula}') from exc
        for token in tokens:
            # Text literals, function names, numbers and operators are never edited.
            if token.type == 'OPERAND' and token.subtype == 'RANGE':
                token.value = self.reference(token.value)
            if token.type == 'FUNC' and token.subtype == 'OPEN' and token.value.upper() in ('INDIRECT(', 'OFFSET('):
                raise ValueError('INDIRECT/OFFSET 动态引用无法可靠地随拆分调整，请先将公式转换为值。')
        return ('=' if had_equals else '') + ''.join(token.value for token in tokens)

    def region(self, region):
        region = CellRange(str(region))
        start, end = (region.min_col, region.max_col) if self.by_columns else (region.min_row, region.max_row)
        mapped = self.interval(start, end)
        if mapped is None:
            return None
        if self.by_columns:
            region.min_col, region.max_col = mapped
        else:
            region.min_row, region.max_row = mapped
        return region.coord


def rebuild_formula(formula, row_mapping, header_rows=0, footer_start=None, sheet_name=None):
    return ReferenceMapper(row_mapping, sheet_name=sheet_name).formula(formula)


def rebuild_formula_by_columns(formula, col_mapping, header_cols=0, footer_start=None, sheet_name=None):
    return ReferenceMapper(col_mapping, True, sheet_name).formula(formula)


def copy_style(source, target):
    if source.has_style:
        for name in ('font', 'fill', 'border', 'alignment', 'protection'):
            setattr(target, name, copy.copy(getattr(source, name)))
        target.number_format = source.number_format


def copy_cell(source, target, mapper):
    target.value = mapper.formula(source.value) if source.data_type == 'f' else source.value
    # Preserve literal strings beginning with '=' and error cells exactly.
    if source.data_type != 'f':
        target.data_type = source.data_type
    copy_style(source, target)
    if source.hyperlink:
        target.hyperlink = copy.copy(source.hyperlink)
        if target.hyperlink.location:
            target.hyperlink.location = mapper.reference(target.hyperlink.location)
    if source.comment:
        target.comment = copy.copy(source.comment)


def validate_features(wb, ws):
    unsupported = []
    if wb.vba_archive is not None and 'xl/vbaProject.bin' in wb.vba_archive.namelist():
        unsupported.append('VBA 宏')
    for name, label in (('_charts', '图表'), ('_images', '图片'), ('_pivots', '数据透视表'), ('tables', 'Excel 表对象')):
        if getattr(ws, name, None):
            unsupported.append(label)
    if unsupported:
        raise ValueError('此工作表含有尚不能完整保留的内容：' + '、'.join(unsupported) + '。请先在副本中处理这些内容后拆分。')


def copy_settings(source, target, mapper):
    for name in ('sheet_format', 'sheet_properties', 'page_margins', 'page_setup', 'print_options', 'views',
                 'protection', 'oddHeader', 'oddFooter', 'evenHeader', 'evenFooter', 'firstHeader', 'firstFooter'):
        setattr(target, name, copy.deepcopy(getattr(source, name)))
    # A single-sheet output must have a visible active sheet and valid view references.
    target.sheet_state = 'visible'
    for view in target.views.sheetView:
        view.tabSelected = True
        view.topLeftCell = None
        if view.pane:
            view.pane = None
        view.selection = []
    from openpyxl.worksheet.views import Selection
    target.sheet_view.selection = [Selection()]
    if source.freeze_panes:
        freeze = source[source.freeze_panes]
        row, col = freeze.row, freeze.column
        if mapper.by_columns:
            col = bisect_left(mapper.indices, col) + 1
        else:
            row = bisect_left(mapper.indices, row) + 1
        target.freeze_panes = f'{get_column_letter(col)}{row}'
    if source.print_area:
        areas = [mapper.region(region) for region in source._print_area.ranges]
        target.print_area = [area for area in areas if area]
    for name in ('print_title_rows', 'print_title_cols'):
        value = getattr(source, name)
        if value:
            result = mapper.reference(value)
            if result != '#REF!':
                setattr(target, name, result)
    if source.auto_filter.ref:
        ref = mapper.region(source.auto_filter.ref)
        if ref:
            target.auto_filter = copy.deepcopy(source.auto_filter)
            target.auto_filter.ref = ref
            if mapper.by_columns:
                old_start = CellRange(source.auto_filter.ref).min_col
                new_start = CellRange(ref).min_col
                filters = []
                for item in target.auto_filter.filterColumn:
                    old_col = old_start + item.colId
                    if old_col in mapper.mapping:
                        item.colId = mapper.mapping[old_col] - new_start
                        filters.append(item)
                target.auto_filter.filterColumn = filters
            if target.auto_filter.sortState:
                state = target.auto_filter.sortState
                state.ref = mapper.region(state.ref) or ref
                conditions = []
                for condition in state.sortCondition:
                    mapped = mapper.region(condition.ref)
                    if mapped:
                        condition.ref = mapped
                        conditions.append(condition)
                state.sortCondition = conditions
    for name, is_columns in (('row_breaks', False), ('col_breaks', True)):
        breaks = copy.deepcopy(getattr(source, name))
        if is_columns == mapper.by_columns:
            kept = []
            for item in breaks.brk:
                index = bisect_right(mapper.indices, item.id)
                if 0 < index < len(mapper.indices):
                    item.id = index
                    if not kept or kept[-1].id != index:
                        kept.append(item)
            breaks.brk = kept
        setattr(target, name, breaks)


def copy_dimensions(source, target, mapper):
    for row, dimension in source.row_dimensions.items():
        new_row = row if mapper.by_columns else mapper.mapping.get(row)
        if new_row is not None:
            result = copy.copy(dimension)
            result.parent = target
            result.index = new_row
            result._style = None
            copy_style(dimension, result)
            target.row_dimensions[new_row] = result
    for letter, dimension in source.column_dimensions.items():
        start = dimension.min or column_index_from_string(letter)
        end = dimension.max or start
        columns = range(start, end + 1)
        for col in columns:
            new_col = mapper.mapping.get(col) if mapper.by_columns else col
            if new_col is None:
                continue
            result = copy.copy(dimension)
            result.parent = target
            result.index = get_column_letter(new_col)
            result.min = result.max = new_col
            result._style = None
            copy_style(dimension, result)
            target.column_dimensions[result.index] = result


def copy_rules(source, target, mapper):
    # Formula rules are relative to each original range's top-left cell.
    from openpyxl.formula.translate import Translator, TranslatorError

    def mapped_rule_formula(formula, old_anchor, retained_anchor):
        if formula is None:
            return None
        text = str(formula)
        try:
            adjusted = Translator('=' + text.lstrip('='), origin=old_anchor).translate_formula(retained_anchor)
        except TranslatorError as exc:
            raise ValueError(f'无法调整条件格式或数据验证公式：{text}') from exc
        result = mapper.formula(adjusted)
        return result if text.startswith('=') else result[1:]

    def anchors(region):
        start = region.min_col if mapper.by_columns else region.min_row
        retained = mapper.indices[bisect_left(mapper.indices, start)]
        old = f'{get_column_letter(region.min_col)}{region.min_row}'
        new = f'{get_column_letter(retained if mapper.by_columns else region.min_col)}{region.min_row if mapper.by_columns else retained}'
        return old, new

    for validation in source.data_validations.dataValidation:
        for region in validation.sqref.ranges:
            mapped = mapper.region(region)
            if not mapped:
                continue
            result = copy.deepcopy(validation)
            result.sqref = mapped
            for name in ('formula1', 'formula2'):
                setattr(result, name, mapped_rule_formula(getattr(validation, name), *anchors(region)))
            target.add_data_validation(result)
    for conditional in source.conditional_formatting:
        for region in conditional.sqref.ranges:
            mapped = mapper.region(region)
            if not mapped:
                continue
            for rule in source.conditional_formatting[conditional]:
                result = copy.deepcopy(rule)
                result.formula = [mapped_rule_formula(f, *anchors(region)) for f in result.formula]
                for visual in (result.colorScale, result.dataBar, result.iconSet):
                    if visual:
                        for threshold in visual.cfvo:
                            if threshold.type == 'formula':
                                threshold.val = mapped_rule_formula(threshold.val, *anchors(region))
                target.conditional_formatting.add(mapped, result)


def build_target(source_ws, group, leading, source_wb, trailing=0, by_columns=False, cell_index=None):
    validate_features(source_wb, source_ws)
    size = source_ws.max_column if by_columns else source_ws.max_row
    included = list(range(1, leading + 1)) + list(group) + list(range(size - trailing + 1, size + 1))
    mapper = ReferenceMapper({old: new for new, old in enumerate(included, 1)}, by_columns, source_ws.title)
    target_wb = openpyxl.Workbook()
    target = target_wb.active
    target.title = source_ws.title
    for name in ('properties', 'security', 'calculation'):
        setattr(target_wb, name, copy.deepcopy(getattr(source_wb, name)))
    target_wb.epoch = source_wb.epoch
    target_wb.loaded_theme = source_wb.loaded_theme
    target_wb.iso_dates = source_wb.iso_dates
    target_wb.template = False
    if target_wb.calculation:
        target_wb.calculation.fullCalcOnLoad = True
        target_wb.calculation.forceFullCalc = True
        target_wb.calculation.calcMode = 'auto'
    copy_settings(source_ws, target, mapper)
    copy_dimensions(source_ws, target, mapper)
    # Visit only populated/styled cells; do not expand sparse sheets into a dense grid.
    cells = source_ws._cells.values() if cell_index is None else chain.from_iterable(cell_index.get(index, ()) for index in included)
    for cell in cells:
        row, col = cell.row, cell.column
        index = col if by_columns else row
        if index not in mapper.mapping or isinstance(cell, MergedCell):
            continue
        new_row, new_col = (row, mapper.mapping[col]) if by_columns else (mapper.mapping[row], col)
        copy_cell(cell, target.cell(new_row, new_col), mapper)
    for region in source_ws.merged_cells.ranges:
        mapped = mapper.region(region)
        if not mapped:
            continue
        new_region = CellRange(mapped)
        copy_cell(source_ws.cell(region.min_row, region.min_col), target.cell(new_region.min_row, new_region.min_col), mapper)
        if new_region.size != {'columns': 1, 'rows': 1}:
            target.merge_cells(mapped)
    copy_rules(source_ws, target, mapper)
    # Keep only names needed by the output; unrelated names may refer to omitted sheets.
    available = {name.name.casefold(): (name, target_wb) for name in source_wb.defined_names.values()}
    available.update({name.name.casefold(): (name, target) for name in source_ws.defined_names.values()})
    processed = set()
    while mapper.names - processed:
        text = next(iter(mapper.names - processed))
        processed.add(text)
        if text.casefold() not in available:
            raise ValueError(f'无法解析公式中的名称：{text}。请先转换为普通单元格引用或值。')
        name, destination = available[text.casefold()]
        result = copy.deepcopy(name)
        result.attr_text = mapper.formula(name.attr_text)
        if destination is target:
            result.localSheetId = 0
        destination.defined_names.add(result)
    return target_wb


def build_target_sheet(source_ws, group_rows, header_rows, source_wb, footer_rows=0):
    return build_target(source_ws, group_rows, header_rows, source_wb, footer_rows)


def build_target_sheet_by_columns(source_ws, group_cols, header_cols, source_wb, footer_cols=0):
    return build_target(source_ws, group_cols, header_cols, source_wb, footer_cols, True)


class SplitFiles(list):
    def __init__(self):
        super().__init__()
        self.warnings = []


def save_groups(source_wb, source_ws, groups, input_path, output_dir, leading, trailing, by_columns=False):
    validate_features(source_wb, source_ws)
    destination = Path(output_dir) if output_dir else input_path.parent / 'split_output'
    destination.mkdir(parents=True, exist_ok=True)
    reserved = {path.name.casefold() for path in destination.iterdir()}
    reserved.add(input_path.name.casefold())
    outputs = SplitFiles()
    cell_index = {}
    for cell in source_ws._cells.values():
        cell_index.setdefault(cell.column if by_columns else cell.row, []).append(cell)
    # Stage the whole batch before publishing; save failures never leave half-written outputs.
    with tempfile.TemporaryDirectory(prefix='.split-', dir=destination) as staging:
        staged = []
        for index, (group_name, indices) in enumerate(groups.items()):
            stem = f'{safe_file_name(input_path.stem)}_{safe_file_name(group_name)}'
            name, suffix = stem + '.xlsx', 2
            while name.casefold() in reserved:
                name = f'{stem}_{suffix}.xlsx'
                suffix += 1
            reserved.add(name.casefold())
            path = destination / name
            workbook = build_target(source_ws, indices, leading, source_wb, trailing, by_columns, cell_index)
            try:
                broken = [cell.coordinate for cell in workbook.active._cells.values()
                          if (cell.data_type == 'f' and '#REF!' in str(cell.value)) or
                          (cell.data_type == 'e' and cell.value == '#REF!')]
                if broken:
                    outputs.warnings.append(f'{name}：{len(broken)} 个单元格含 #REF!（如 {"、".join(broken[:5])}），请核对被剔除数据的引用。')
                temp_path = Path(staging) / f'{index}.xlsx'
                workbook.save(temp_path)
            finally:
                workbook.close()
            staged.append((temp_path, path))
        try:
            for temp_path, path in staged:
                # Exclusive creation also protects against another run choosing this name.
                with path.open('xb') as output:
                    outputs.append(path)
                    with temp_path.open('rb') as source:
                        import shutil
                        shutil.copyfileobj(source, output)
                    output.flush()
                    os.fsync(output.fileno())
        except Exception:
            for path in outputs:
                path.unlink(missing_ok=True)
            raise
    return outputs
