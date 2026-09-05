from __future__ import annotations

import argparse
import queue
import threading
import sys
import warnings
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import tkinter as tk
from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox, ttk
from split_core import (
    normalize_group_value, safe_file_name, collect_groups, collect_group_rows,
    collect_group_columns, validate_parameters, save_groups, build_target_sheet,
    build_target_sheet_by_columns, rebuild_formula, rebuild_formula_by_columns,
)


OPENPYXL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
CONVERTIBLE_EXTENSIONS = {".xls", ".et"}


def parse_args(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="按关键列或关键行拆分 Excel，并保留原始表格样式和格式。"
    )
    parser.add_argument("--input", help="待拆分的 Excel 文件路径")
    parser.add_argument("--sheet", help="要拆分的工作表名称，不传则默认第一个工作表")
    parser.add_argument(
        "--mode",
        choices=["column", "row"],
        default="column",
        help="拆分方式：column=按关键列拆分数据行（默认），row=按关键行拆分数据列",
    )
    parser.add_argument("--header-rows", type=int, help="表头行数（column 模式）")
    parser.add_argument("--footer-rows", type=int, default=0, help="表尾行数（固定在每个文件末尾，column 模式），默认 0")
    parser.add_argument("--key-column", type=int, help="关键列序号，从 1 开始（column 模式）")
    parser.add_argument("--header-cols", type=int, help="左侧固定列数（row 模式）")
    parser.add_argument("--footer-cols", type=int, default=0, help="右侧固定列数（固定在每个文件右侧，row 模式），默认 0")
    parser.add_argument("--key-row", type=int, help="关键行序号，从 1 开始（row 模式）")
    parser.add_argument("--output-dir", help="输出目录，默认在源文件同级目录生成 split_output")
    args = parser.parse_args(argv)
    supplied = sys.argv[1:] if argv is None else argv
    if supplied:
        if not args.input:
            parser.error("命令行运行必须提供 --input。")
        required = ("header_cols", "key_row") if args.mode == "row" else ("header_rows", "key_column")
        for name in required:
            if getattr(args, name) is None:
                parser.error(f"缺少参数 --{name.replace('_', '-')}。")
    return args


def load_workbook_compatible(input_path: Path) -> Tuple[openpyxl.Workbook, Optional[tempfile.TemporaryDirectory]]:
    input_path = Path(input_path)
    if not input_path.is_file():
        raise ValueError(f"找不到输入文件：{input_path}")
    suffix = input_path.suffix.lower()
    if suffix in OPENPYXL_EXTENSIONS:
        return read_workbook(input_path, keep_vba=suffix in {".xlsm", ".xltm"}), None

    if suffix == ".et":
        from zipfile import is_zipfile
        if is_zipfile(input_path):
            with input_path.open("rb") as file_handle:
                return read_workbook(file_handle, keep_vba=True), None

    if suffix in CONVERTIBLE_EXTENSIONS:
        converted_path, temp_dir = convert_to_xlsx(input_path)
        try:
            return read_workbook(converted_path), temp_dir
        except Exception:
            temp_dir.cleanup()
            raise

    raise ValueError(f"暂不支持该文件格式：{suffix}")


def read_workbook(path, **options):
    with warnings.catch_warnings(record=True) as captured:
        warnings.simplefilter("always", UserWarning)
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=False, rich_text=True, **options)
    lost_features = [str(item.message) for item in captured if "removed" in str(item.message) or "not supported" in str(item.message)]
    if lost_features:
        close_workbook_compatible(workbook, None)
        raise ValueError("文件包含无法完整保留的功能：" + "；".join(lost_features))
    return workbook


def select_sheet(workbook, name=None):
    if not workbook.worksheets:
        raise ValueError("文件中没有可拆分的数据工作表。")
    if name is None:
        return workbook.worksheets[0]
    for sheet in workbook.worksheets:
        if sheet.title == name:
            return sheet
    raise ValueError(f"找不到工作表：{name}")


def close_workbook_compatible(workbook: openpyxl.Workbook, temp_dir: Optional[tempfile.TemporaryDirectory]) -> None:
    try:
        workbook.close()
        if workbook.vba_archive is not None:
            workbook.vba_archive.close()
    finally:
        if temp_dir:
            temp_dir.cleanup()


def convert_to_xlsx(input_path: Path) -> Tuple[Path, tempfile.TemporaryDirectory]:
    temp_dir = tempfile.TemporaryDirectory()
    temp_path = Path(temp_dir.name)

    try:
        try:
            converted_path = convert_with_libreoffice(input_path, temp_path)
        except (OSError, subprocess.SubprocessError):
            converted_path = None
        if converted_path:
            return converted_path, temp_dir

        converted_path = convert_with_windows_com(input_path, temp_path)
        if converted_path:
            return converted_path, temp_dir
    except Exception:
        temp_dir.cleanup()
        raise

    temp_dir.cleanup()
    raise ValueError(
        "无法自动转换该文件。请安装 LibreOffice，或在 Windows 上安装 Excel/WPS 后重试。"
    )


def convert_with_libreoffice(input_path: Path, output_dir: Path) -> Optional[Path]:
    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        return None

    result = subprocess.run(
        [
            executable,
            f"-env:UserInstallation={(output_dir / 'lo-profile').resolve().as_uri()}",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(input_path.resolve()),
        ],
        capture_output=True,
        text=True,
        timeout=120,
        check=False,
        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
    )
    if result.returncode != 0:
        return None

    expected_path = output_dir / f"{input_path.stem}.xlsx"
    if expected_path.exists():
        return expected_path

    converted_files = list(output_dir.glob("*.xlsx"))
    return converted_files[0] if converted_files else None


def convert_with_windows_com(input_path: Path, output_dir: Path) -> Optional[Path]:
    if not shutil.which("powershell"):
        return None

    output_path = output_dir / f"{input_path.stem}.xlsx"
    script_path = output_dir / "convert_spreadsheet.ps1"
    script_path.write_text(
        r"""
param(
    [string]$InputPath,
    [string]$OutputPath
)

$ErrorActionPreference = "Stop"
$progIds = @("Excel.Application", "Ket.Application", "ET.Application")
$app = $null

foreach ($progId in $progIds) {
    try {
        $app = New-Object -ComObject $progId
        break
    } catch {
        $app = $null
    }
}

if ($null -eq $app) {
    throw "未找到可用的 Excel/WPS COM 转换器"
}

$app.Visible = $false
$app.DisplayAlerts = $false
$workbook = $null

try {
    $app.AutomationSecurity = 3
    $workbook = $app.Workbooks.Open($InputPath, 0, $true)
    if ($workbook.HasVBProject) {
        throw "文件包含 VBA 宏，转换为 xlsx 会丢失宏。请先使用不含宏的副本。"
    }
    $workbook.SaveAs($OutputPath, 51)
} finally {
    if ($null -ne $workbook) {
        $workbook.Close($false)
    }
    $app.Quit()
}
""",
        encoding="utf-8-sig",
    )
    result = subprocess.run(
        [
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script_path),
            str(input_path.resolve()),
            str(output_path),
        ],
        capture_output=True,
        text=True,
        timeout=120,
        check=False,
        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
    )
    if result.returncode != 0 or not output_path.exists():
        return None
    return output_path


def _split_workbook(input_path, sheet_name, leading, key, output_dir, trailing=0, by_columns=False):
    input_path = Path(input_path)
    source_wb, temp_dir = load_workbook_compatible(input_path)
    try:
        source_ws = select_sheet(source_wb, sheet_name)
        groups = collect_groups(source_ws, leading, key, trailing, by_columns)
        return save_groups(source_wb, source_ws, groups, input_path, output_dir, leading, trailing, by_columns)
    finally:
        close_workbook_compatible(source_wb, temp_dir)


def split_workbook(input_path, sheet_name, header_rows, key_column, output_dir=None, footer_rows=0):
    return _split_workbook(input_path, sheet_name, header_rows, key_column, output_dir, footer_rows)


def split_workbook_by_row(input_path, sheet_name, header_cols, key_row, output_dir=None, footer_cols=0):
    return _split_workbook(input_path, sheet_name, header_cols, key_row, output_dir, footer_cols, True)


def preview_indices(total, *focus):
    indices = set(range(1, min(total, 20) + 1))
    indices.update(range(max(1, total - 2), total + 1))
    for value in focus:
        if value is not None:
            indices.update(range(max(1, value - 3), min(total, value + 4) + 1))
    return sorted(indices)


def load_workbook_info(input_path: Path) -> Dict[str, object]:
    workbook, temp_dir = load_workbook_compatible(input_path)
    try:
        sheet_names = [sheet.title for sheet in workbook.worksheets]
        first_sheet = select_sheet(workbook)
        column_preview = []
        preview_row = min(first_sheet.max_row, 6)
        for column in preview_indices(first_sheet.max_column):
            value = first_sheet.cell(preview_row, column).value
            display = normalize_group_value(value) or "(空)"
            column_preview.append(f"{column}. {get_column_letter(column)} - {display}")
        return {
            "sheet_names": sheet_names,
            "max_row": first_sheet.max_row,
            "max_column": first_sheet.max_column,
            "column_preview": column_preview,
        }
    finally:
        close_workbook_compatible(workbook, temp_dir)


def build_sheet_preview(
    input_path: Path,
    sheet_name: str,
    header_rows: int,
    key_column: int,
    footer_rows: int = 0,
) -> Dict[str, object]:
    workbook, temp_dir = load_workbook_compatible(input_path)
    try:
        ws = select_sheet(workbook, sheet_name)
        validate_parameters(ws, header_rows, key_column, footer_rows)
        safe_header_rows, safe_key_column, safe_footer_rows = header_rows, key_column, footer_rows
        footer_start = ws.max_row - safe_footer_rows + 1 if safe_footer_rows else None

        preview_columns = preview_indices(ws.max_column, safe_key_column)

        column_headers: List[str] = []
        header_reference_row = max(1, min(ws.max_row, safe_header_rows))
        for column in preview_columns:
            header_text = normalize_group_value(ws.cell(row=header_reference_row, column=column).value) or "(空)"
            prefix = "[关键] " if column == safe_key_column else ""
            column_headers.append(f"{prefix}{get_column_letter(column)}列/{column}: {header_text}")

        def row_values(target_row: int) -> List[str]:
            collected = []
            for column in preview_columns:
                value = normalize_group_value(ws.cell(row=target_row, column=column).value) or ""
                collected.append(value.replace("\n", " "))
            return collected

        preview_rows: List[Dict[str, object]] = []
        preview_data_limit = min(ws.max_row, safe_header_rows + 8)
        if footer_start is not None:
            preview_data_limit = min(preview_data_limit, footer_start - 1)
        for row in preview_indices(preview_data_limit, safe_header_rows):
            row_kind = "header" if row <= safe_header_rows else "data"
            preview_rows.append({"row_no": row, "kind": row_kind, "values": row_values(row)})
            if row == safe_header_rows and row < ws.max_row:
                preview_rows.append(
                    {
                        "row_no": "↓",
                        "kind": "split",
                        "values": ["以下数据行将参与拆分"] + [""] * (len(preview_columns) - 1),
                    }
                )
        if footer_start is not None:
            preview_rows.append(
                {
                    "row_no": "↓",
                    "kind": "split",
                    "values": ["以下为固定表尾，会附加到每个文件末尾"] + [""] * (len(preview_columns) - 1),
                }
            )
            for row in preview_indices(ws.max_row, footer_start):
                if row < footer_start:
                    continue
                preview_rows.append({"row_no": row, "kind": "footer", "values": row_values(row)})

        key_column_preview: List[str] = []
        for column in preview_columns:
            value = normalize_group_value(ws.cell(header_reference_row, column).value) or "(空)"
            marker = " <关键列>" if column == safe_key_column else ""
            key_column_preview.append(f"{column}. {get_column_letter(column)} - {value}{marker}")

        groups = collect_group_rows(ws, safe_header_rows, safe_key_column, safe_footer_rows)
        split_objects = [{"name": name, "count": len(rows)} for name, rows in groups.items()]

        return {
            "sheet_title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "preview_columns": preview_columns,
            "column_headers": column_headers,
            "preview_rows": preview_rows,
            "key_column_preview": key_column_preview,
            "split_objects": split_objects,
            "group_count": len(split_objects),
        }
    finally:
        close_workbook_compatible(workbook, temp_dir)


def build_sheet_preview_by_row(
    input_path: Path,
    sheet_name: str,
    header_cols: int,
    key_row: int,
    footer_cols: int = 0,
) -> Dict[str, object]:
    workbook, temp_dir = load_workbook_compatible(input_path)
    try:
        ws = select_sheet(workbook, sheet_name)
        validate_parameters(ws, header_cols, key_row, footer_cols, True)
        safe_header_cols, safe_key_row, safe_footer_cols = header_cols, key_row, footer_cols
        footer_start = ws.max_column - safe_footer_cols + 1 if safe_footer_cols else None

        preview_columns = preview_indices(ws.max_column, safe_header_cols, footer_start)

        column_headers: List[str] = []
        for column in preview_columns:
            key_text = normalize_group_value(ws.cell(row=safe_key_row, column=column).value) or "(空)"
            if column <= safe_header_cols:
                prefix = "[左固定] "
            elif footer_start is not None and column >= footer_start:
                prefix = "[右固定] "
            else:
                prefix = ""
            column_headers.append(f"{prefix}{get_column_letter(column)}列/{column}: {key_text}")

        def row_values(target_row: int) -> List[str]:
            collected = []
            for column in preview_columns:
                value = normalize_group_value(ws.cell(row=target_row, column=column).value) or ""
                collected.append(value.replace("\n", " "))
            return collected

        preview_rows: List[Dict[str, object]] = []
        preview_limit = min(ws.max_row, max(safe_key_row, 1) + 8)
        for row in preview_indices(preview_limit, safe_key_row):
            row_kind = "keyrow" if row == safe_key_row else "data"
            row_label = f"{row} <关键行>" if row == safe_key_row else row
            preview_rows.append({"row_no": row_label, "kind": row_kind, "values": row_values(row)})
        if preview_limit < ws.max_row:
            preview_rows.append(
                {
                    "row_no": "…",
                    "kind": "split",
                    "values": [f"（其余 {ws.max_row - preview_limit} 行未显示，所有行都会完整保留）"]
                    + [""] * (len(preview_columns) - 1),
                }
            )

        key_row_preview: List[str] = []
        preview_row_limit = min(ws.max_row, 30)
        for row in sorted(set(range(1, preview_row_limit + 1)) | {safe_key_row}):
            samples: List[str] = []
            for column in preview_columns:
                value = normalize_group_value(ws.cell(row=row, column=column).value)
                if value:
                    samples.append(value.replace("\n", " "))
                if len(samples) >= 3:
                    break
            display = " | ".join(samples) if samples else "(空行)"
            marker = " <关键行>" if row == safe_key_row else ""
            key_row_preview.append(f"{row}. {display}{marker}")

        groups = collect_group_columns(ws, safe_header_cols, safe_key_row, safe_footer_cols)
        split_objects = [{"name": name, "count": len(cols)} for name, cols in groups.items()]

        return {
            "sheet_title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "preview_columns": preview_columns,
            "column_headers": column_headers,
            "preview_rows": preview_rows,
            "key_row_preview": key_row_preview,
            "split_objects": split_objects,
            "group_count": len(split_objects),
        }
    finally:
        close_workbook_compatible(workbook, temp_dir)


class SplitterApp:
    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title("Excel 拆表工具")
        self.root.geometry("980x760")
        self.root.minsize(900, 660)
        self.root.configure(bg="#F4F6FA")

        self.input_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.mode_var = tk.StringVar(value="column")
        self.header_rows_var = tk.StringVar(value="6")
        self.footer_rows_var = tk.StringVar(value="0")
        self.key_column_var = tk.StringVar(value="3")
        self.header_cols_var = tk.StringVar(value="1")
        self.footer_cols_var = tk.StringVar(value="0")
        self.key_row_var = tk.StringVar(value="1")
        self.output_var = tk.StringVar()
        self.summary_var = tk.StringVar(value="先选择 Excel 文件，界面会自动加载工作表和列信息。")
        self.preview_column_map: Dict[str, int] = {}

        self._jobs = queue.Queue()
        self._results = queue.Queue()
        self._versions = {}
        self._preview_after = None
        self._splitting = False
        self._closed = False
        self._build_ui()
        threading.Thread(target=self._worker, daemon=True).start()
        self._poll_after = self.root.after(75, self._poll_jobs)
        self.root.protocol("WM_DELETE_WINDOW", self._close)
        self.header_rows_var.trace_add("write", self.on_option_changed)
        self.footer_rows_var.trace_add("write", self.on_option_changed)
        self.key_column_var.trace_add("write", self.on_option_changed)
        self.header_cols_var.trace_add("write", self.on_option_changed)
        self.footer_cols_var.trace_add("write", self.on_option_changed)
        self.key_row_var.trace_add("write", self.on_option_changed)

    def _build_ui(self) -> None:
        # ---- 配色与字体 ----
        BG = "#F4F6FA"            # 窗口底色
        ACCENT = "#2F6FED"        # 主色（蓝）
        ACCENT_HOVER = "#255FCC"
        ACCENT_PRESS = "#1E4FB0"
        TEXT = "#1F2937"          # 正文
        MUTED = "#6B7280"         # 辅助灰
        BORDER = "#D9DEE7"        # 边框
        FONT = "Microsoft YaHei UI"

        style = ttk.Style()
        style.theme_use("clam")

        style.configure(".", font=(FONT, 10), background=BG, foreground=TEXT, bordercolor=BORDER)
        style.configure("TFrame", background=BG)
        style.configure("TLabel", background=BG, foreground=TEXT)
        style.configure("Title.TLabel", font=(FONT, 17, "bold"), background=BG, foreground=TEXT)
        style.configure("Muted.TLabel", background=BG, foreground=MUTED)
        style.configure("Summary.TLabel", background=BG, foreground=ACCENT, font=(FONT, 10, "bold"))
        style.configure(
            "TLabelframe", background=BG, bordercolor=BORDER, relief="solid", borderwidth=1
        )
        style.configure(
            "TLabelframe.Label", background=BG, foreground=TEXT, font=(FONT, 10, "bold")
        )
        style.configure("TRadiobutton", background=BG, foreground=TEXT)
        style.map("TRadiobutton", background=[("active", BG)])
        style.configure(
            "TButton",
            padding=(14, 6),
            background="#FFFFFF",
            foreground=TEXT,
            bordercolor=BORDER,
            focuscolor=ACCENT,
        )
        style.map(
            "TButton",
            background=[("pressed", "#E5EAF3"), ("active", "#EFF3FB")],
            bordercolor=[("active", ACCENT)],
        )
        style.configure(
            "Accent.TButton",
            padding=(22, 9),
            font=(FONT, 11, "bold"),
            background=ACCENT,
            foreground="#FFFFFF",
            bordercolor=ACCENT,
        )
        style.map(
            "Accent.TButton",
            background=[("pressed", ACCENT_PRESS), ("active", ACCENT_HOVER)],
            foreground=[("disabled", "#FFFFFF")],
        )
        style.configure(
            "TEntry", padding=(6, 4), fieldbackground="#FFFFFF",
            bordercolor=BORDER, lightcolor=BORDER, darkcolor=BORDER,
        )
        style.configure(
            "TCombobox", padding=(6, 4), fieldbackground="#FFFFFF", bordercolor=BORDER
        )
        style.map(
            "TCombobox",
            fieldbackground=[("readonly", "#FFFFFF")],
            bordercolor=[("active", ACCENT)],
        )
        style.configure(
            "Treeview",
            background="#FFFFFF",
            fieldbackground="#FFFFFF",
            foreground=TEXT,
            rowheight=24,
            bordercolor=BORDER,
            font=(FONT, 9),
        )
        style.configure(
            "Treeview.Heading",
            font=(FONT, 9, "bold"),
            background="#EEF2F8",
            foreground=TEXT,
            padding=(6, 4),
        )
        style.map(
            "Treeview",
            background=[("selected", "#D6E4FF")],
            foreground=[("selected", TEXT)],
        )
        style.configure("TScrollbar", background="#E5EAF3", troughcolor=BG, bordercolor=BG, arrowcolor=MUTED)
        style.configure("TSeparator", background=BORDER)

        container = ttk.Frame(self.root, padding=18)
        container.pack(fill="both", expand=True)
        container.columnconfigure(1, weight=1)
        container.rowconfigure(6, weight=1)

        title = ttk.Label(container, text="Excel 拆表工具", style="Title.TLabel")
        title.grid(row=0, column=0, columnspan=3, sticky="w")

        desc = ttk.Label(
            container,
            text="支持按关键列（拆数据行）或按关键行（拆数据列）两种方式拆分。会保留原表样式、格式和公式。",
            style="Muted.TLabel",
        )
        desc.grid(row=1, column=0, columnspan=3, sticky="w", pady=(4, 6))

        ttk.Separator(container, orient="horizontal").grid(
            row=1, column=0, columnspan=3, sticky="sew", pady=(0, 10)
        )

        ttk.Label(container, text="Excel 文件").grid(row=2, column=0, sticky="w", pady=6)
        ttk.Entry(container, textvariable=self.input_var).grid(row=2, column=1, sticky="ew", pady=6, padx=(0, 8))
        ttk.Button(container, text="浏览", command=self.choose_input).grid(row=2, column=2, sticky="ew", pady=6)

        ttk.Label(container, text="工作表").grid(row=3, column=0, sticky="w", pady=6)
        self.sheet_combo = ttk.Combobox(container, textvariable=self.sheet_var, state="readonly")
        self.sheet_combo.grid(row=3, column=1, sticky="ew", pady=6, padx=(0, 8))
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_changed)
        ttk.Button(container, text="刷新", command=self.reload_workbook_info).grid(row=3, column=2, sticky="ew", pady=6)

        options = ttk.Frame(container)
        options.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(8, 12))
        options.columnconfigure(0, weight=1)

        mode_bar = ttk.Frame(options)
        mode_bar.grid(row=0, column=0, sticky="w", pady=(0, 8))
        ttk.Label(mode_bar, text="拆分方式").grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(
            mode_bar,
            text="按关键列拆分（每组数据行一个文件）",
            variable=self.mode_var,
            value="column",
            command=self.on_mode_changed,
        ).grid(row=0, column=1, sticky="w", padx=(12, 0))
        ttk.Radiobutton(
            mode_bar,
            text="按关键行拆分（每组数据列一个文件）",
            variable=self.mode_var,
            value="row",
            command=self.on_mode_changed,
        ).grid(row=0, column=2, sticky="w", padx=(16, 0))

        self.column_options = ttk.Frame(options)
        self.column_options.grid(row=1, column=0, sticky="ew")
        self.column_options.columnconfigure(1, weight=1)
        self.column_options.columnconfigure(3, weight=1)

        ttk.Label(self.column_options, text="表头行数").grid(row=0, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(self.column_options, textvariable=self.header_rows_var, width=10).grid(
            row=0, column=1, sticky="w", padx=(8, 24), pady=(0, 6)
        )
        ttk.Label(self.column_options, text="表尾行数").grid(row=0, column=2, sticky="w", pady=(0, 6))
        ttk.Entry(self.column_options, textvariable=self.footer_rows_var, width=10).grid(
            row=0, column=3, sticky="w", padx=(8, 0), pady=(0, 6)
        )

        ttk.Label(self.column_options, text="关键列序号").grid(row=1, column=0, sticky="w")
        ttk.Entry(self.column_options, textvariable=self.key_column_var, width=10).grid(
            row=1, column=1, sticky="w", padx=(8, 24)
        )
        ttk.Label(self.column_options, text="或点选列").grid(row=1, column=2, sticky="w")
        self.key_column_combo = ttk.Combobox(self.column_options, state="readonly")
        self.key_column_combo.grid(row=1, column=3, sticky="ew", padx=(8, 0))
        self.key_column_combo.bind("<<ComboboxSelected>>", self.on_key_column_selected)

        self.row_options = ttk.Frame(options)
        self.row_options.grid(row=1, column=0, sticky="ew")
        self.row_options.columnconfigure(1, weight=1)
        self.row_options.columnconfigure(3, weight=1)

        ttk.Label(self.row_options, text="左侧固定列数").grid(row=0, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(self.row_options, textvariable=self.header_cols_var, width=10).grid(
            row=0, column=1, sticky="w", padx=(8, 24), pady=(0, 6)
        )
        ttk.Label(self.row_options, text="右侧固定列数").grid(row=0, column=2, sticky="w", pady=(0, 6))
        ttk.Entry(self.row_options, textvariable=self.footer_cols_var, width=10).grid(
            row=0, column=3, sticky="w", padx=(8, 0), pady=(0, 6)
        )

        ttk.Label(self.row_options, text="关键行序号").grid(row=1, column=0, sticky="w")
        ttk.Entry(self.row_options, textvariable=self.key_row_var, width=10).grid(
            row=1, column=1, sticky="w", padx=(8, 24)
        )
        ttk.Label(self.row_options, text="或点选行").grid(row=1, column=2, sticky="w")
        self.key_row_combo = ttk.Combobox(self.row_options, state="readonly")
        self.key_row_combo.grid(row=1, column=3, sticky="ew", padx=(8, 0))
        self.key_row_combo.bind("<<ComboboxSelected>>", self.on_key_row_selected)

        self.row_options.grid_remove()

        summary = ttk.Label(container, textvariable=self.summary_var, style="Summary.TLabel", wraplength=900)
        summary.grid(row=5, column=0, columnspan=3, sticky="new", pady=(0, 12))

        preview_area = ttk.Frame(container)
        preview_area.grid(row=6, column=0, columnspan=3, sticky="nsew", pady=(0, 12))
        preview_area.columnconfigure(0, weight=5)
        preview_area.columnconfigure(1, weight=2)
        preview_area.rowconfigure(0, weight=1)

        preview_box = ttk.LabelFrame(preview_area, text="表格预览", height=520)
        preview_box.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        preview_box.columnconfigure(0, weight=1)
        preview_box.rowconfigure(0, weight=1)
        preview_box.rowconfigure(1, weight=1)
        preview_box.grid_propagate(False)
        self.preview_hint_var = tk.StringVar(value="浅黄色为表头，浅绿色为固定表尾，蓝色为分界提示；关键列会在列标题中标记，可用底部横向滚动条左右查看。")
        ttk.Label(preview_box, textvariable=self.preview_hint_var, style="Muted.TLabel").grid(
            row=0, column=0, sticky="w", pady=(0, 8)
        )
        preview_table_frame = ttk.Frame(preview_box)
        preview_table_frame.grid(row=1, column=0, sticky="nsew")
        preview_table_frame.columnconfigure(0, weight=1)
        preview_table_frame.rowconfigure(0, weight=1)
        self.preview_table = ttk.Treeview(preview_table_frame, show="headings")
        self.preview_table.grid(row=0, column=0, sticky="nsew")
        self.preview_table.bind("<ButtonRelease-1>", self.on_preview_table_click)
        preview_scroll_y = ttk.Scrollbar(preview_table_frame, orient="vertical", command=self.preview_table.yview)
        preview_scroll_y.grid(row=0, column=1, sticky="ns")
        preview_scroll_x = ttk.Scrollbar(preview_table_frame, orient="horizontal", command=self.preview_table.xview)
        preview_scroll_x.grid(row=1, column=0, sticky="ew")
        self.preview_table.configure(yscrollcommand=preview_scroll_y.set, xscrollcommand=preview_scroll_x.set)
        self.preview_table.tag_configure("header", background="#fff4cc")
        self.preview_table.tag_configure("split", background="#dceeff")
        self.preview_table.tag_configure("footer", background="#e2efda")
        self.preview_table.tag_configure("keyrow", background="#ffe0e0")

        side_panel = ttk.Frame(preview_area)
        side_panel.grid(row=0, column=1, sticky="nsew")
        side_panel.columnconfigure(0, weight=1)
        side_panel.rowconfigure(0, weight=1)
        side_panel.rowconfigure(1, weight=1)

        self.reference_box = ttk.LabelFrame(side_panel, text="关键列参考", height=220)
        column_box = self.reference_box
        column_box.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
        column_box.columnconfigure(0, weight=1)
        column_box.rowconfigure(0, weight=1)
        column_box.grid_propagate(False)
        self.preview_text = tk.Text(
            column_box,
            height=8,
            wrap="word",
            font=("Microsoft YaHei UI", 9),
            bg="#FFFFFF",
            fg="#1F2937",
            relief="flat",
            highlightthickness=1,
            highlightbackground="#D9DEE7",
            highlightcolor="#2F6FED",
            padx=8,
            pady=6,
        )
        self.preview_text.grid(row=0, column=0, sticky="nsew")
        self.preview_text.configure(state="disabled")

        self.split_box = ttk.LabelFrame(side_panel, text="拆分对象预览", height=292)
        split_box = self.split_box
        split_box.grid(row=1, column=0, sticky="nsew")
        split_box.columnconfigure(0, weight=1)
        split_box.rowconfigure(0, weight=1)
        split_box.grid_propagate(False)
        self.split_preview_table = ttk.Treeview(split_box, columns=("name", "count"), show="headings", height=10)
        self.split_preview_table.heading("name", text="拆分对象")
        self.split_preview_table.heading("count", text="行数")
        self.split_preview_table.column("name", width=180, anchor="w")
        self.split_preview_table.column("count", width=70, anchor="center")
        self.split_preview_table.grid(row=0, column=0, sticky="nsew")
        split_scroll_y = ttk.Scrollbar(split_box, orient="vertical", command=self.split_preview_table.yview)
        split_scroll_y.grid(row=0, column=1, sticky="ns")
        self.split_preview_table.configure(yscrollcommand=split_scroll_y.set)

        ttk.Label(container, text="输出目录").grid(row=7, column=0, sticky="nw", pady=6)
        output_frame = ttk.Frame(container)
        output_frame.grid(row=7, column=1, columnspan=2, sticky="ew", pady=6)
        output_frame.columnconfigure(0, weight=1)
        ttk.Entry(output_frame, textvariable=self.output_var).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ttk.Button(output_frame, text="选择目录", command=self.choose_output_dir).grid(row=0, column=1, sticky="ew")

        action_bar = ttk.Frame(container)
        action_bar.grid(row=8, column=0, columnspan=3, sticky="ew", pady=(16, 0))
        action_bar.columnconfigure(0, weight=1)
        ttk.Label(
            action_bar,
            text="输出为空时，默认生成到源文件同级目录的 split_output。",
            style="Muted.TLabel",
        ).grid(row=0, column=0, sticky="w")
        ttk.Button(
            action_bar, text="开始拆分", command=self.run_split, style="Accent.TButton"
        ).grid(row=0, column=1, sticky="e")

    def choose_input(self) -> None:
        path = filedialog.askopenfilename(
            title="选择待拆分的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm *.xls *.et")],
        )
        if not path:
            return
        self.input_var.set(path)
        if not self.output_var.get().strip():
            self.output_var.set(str(Path(path).parent / "split_output"))
        self.reload_workbook_info()

    def choose_output_dir(self) -> None:
        initial_dir = self.output_var.get().strip() or str(Path(self.input_var.get().strip()).parent)
        path = filedialog.askdirectory(title="选择输出目录", initialdir=initial_dir, mustexist=False)
        if path:
            self.output_var.set(path)

    def _worker(self):
        while True:
            job = self._jobs.get()
            if job is None or self._closed:
                return
            kind, version, work, done = job
            if self._versions.get(kind) != version:
                continue
            try:
                value, error = work(), None
            except Exception as exc:
                value, error = None, exc
            self._results.put((kind, version, done, value, error))

    def _start_job(self, kind, work, done):
        version = self._versions.get(kind, 0) + 1
        self._versions[kind] = version
        self._jobs.put((kind, version, work, done))

    def _poll_jobs(self):
        try:
            while True:
                kind, version, done, value, error = self._results.get_nowait()
                if self._versions.get(kind) == version:
                    done(value, error)
        except queue.Empty:
            pass
        finally:
            if not self._closed:
                self._poll_after = self.root.after(75, self._poll_jobs)

    def _close(self):
        if self._splitting:
            messagebox.showinfo("正在拆分", "请等待当前拆分完成后再关闭窗口。")
            return
        self._closed = True
        self.root.after_cancel(self._poll_after)
        if self._preview_after is not None:
            self.root.after_cancel(self._preview_after)
        self._jobs.put(None)
        self.root.destroy()

    def reload_workbook_info(self) -> None:
        input_path = self.input_var.get().strip()
        if not input_path or self._splitting:
            return
        self._versions["preview"] = self._versions.get("preview", 0) + 1
        self.summary_var.set("正在读取文件……")
        self.clear_preview_table()
        self.clear_split_preview_table()

        def done(info, error):
            if self.input_var.get().strip() != input_path:
                return
            if error:
                self.sheet_var.set("")
                self.sheet_combo["values"] = ()
                self.summary_var.set(f"读取失败：{error}")
                messagebox.showerror("读取失败", str(error))
                return
            sheet_names = info["sheet_names"]
            self.sheet_combo["values"] = sheet_names
            if self.sheet_var.get().strip() not in sheet_names:
                self.sheet_var.set(sheet_names[0])
            self.refresh_sheet_preview()

        self._start_job("load", lambda: load_workbook_info(Path(input_path)), done)

    def on_sheet_changed(self, _event: object) -> None:
        self.refresh_sheet_preview()

    def on_option_changed(self, *_args: object) -> None:
        self._versions["preview"] = self._versions.get("preview", 0) + 1
        if self._preview_after is not None:
            self.root.after_cancel(self._preview_after)
        self._preview_after = self.root.after(300, self.refresh_sheet_preview)

    def on_key_column_selected(self, _event: object) -> None:
        selected = self.key_column_combo.get().strip()
        if not selected:
            return
        try:
            column_no = int(selected.split(".", 1)[0])
        except ValueError:
            return
        if self.key_column_var.get().strip() != str(column_no):
            self.key_column_var.set(str(column_no))

    def on_key_row_selected(self, _event: object) -> None:
        selected = self.key_row_combo.get().strip()
        if not selected:
            return
        try:
            row_no = int(selected.split(".", 1)[0])
        except ValueError:
            return
        if self.key_row_var.get().strip() != str(row_no):
            self.key_row_var.set(str(row_no))

    def on_mode_changed(self) -> None:
        mode = self.mode_var.get()
        if mode == "row":
            self.column_options.grid_remove()
            self.row_options.grid()
            self.reference_box.configure(text="关键行参考")
            self.split_box.configure(text="拆分对象预览")
            self.split_preview_table.heading("count", text="列数")
            self.preview_hint_var.set(
                "浅红色为关键行，列标题中会标记左右固定列；点击表格中的任意一行即可将其设为关键行。"
            )
        else:
            self.row_options.grid_remove()
            self.column_options.grid()
            self.reference_box.configure(text="关键列参考")
            self.split_box.configure(text="拆分对象预览")
            self.split_preview_table.heading("count", text="行数")
            self.preview_hint_var.set(
                "浅黄色为表头，浅绿色为固定表尾，蓝色为分界提示；关键列会在列标题中标记，可用底部横向滚动条左右查看。"
            )
        self.refresh_sheet_preview()

    def on_preview_table_click(self, event: tk.Event) -> None:
        region = self.preview_table.identify("region", event.x, event.y)
        if self.mode_var.get() == "row":
            if region != "cell":
                return
            item_id = self.preview_table.identify_row(event.y)
            if not item_id:
                return
            values = self.preview_table.item(item_id, "values")
            if not values:
                return
            row_text = str(values[0]).split(" ", 1)[0].strip()
            if not row_text.isdigit():
                return
            if self.key_row_var.get().strip() != row_text:
                self.key_row_var.set(row_text)
            return

        if region != "heading":
            return
        column_id = self.preview_table.identify_column(event.x)
        if not column_id or column_id == "#1":
            return
        mapped = self.preview_column_map.get(column_id)
        if mapped is not None and self.key_column_var.get().strip() != str(mapped):
            self.key_column_var.set(str(mapped))

    def refresh_sheet_preview(self) -> None:
        if self._preview_after is not None:
            self.root.after_cancel(self._preview_after)
            self._preview_after = None
        if self._splitting:
            return
        input_path = self.input_var.get().strip()
        sheet_name = self.sheet_var.get().strip()
        if not input_path or not sheet_name:
            return
        self._versions["preview"] = self._versions.get("preview", 0) + 1
        mode = self.mode_var.get()
        try:
            if mode == "row":
                leading, trailing, key = map(int, (self.header_cols_var.get(), self.footer_cols_var.get(), self.key_row_var.get()))
                preview_function = build_sheet_preview_by_row
            else:
                leading, trailing, key = map(int, (self.header_rows_var.get(), self.footer_rows_var.get(), self.key_column_var.get()))
                preview_function = build_sheet_preview
        except ValueError:
            self._preview_error(ValueError("拆分参数必须填写整数。"))
            return
        self.summary_var.set("正在生成预览……")

        def done(info, error):
            if self.input_var.get().strip() != input_path:
                return
            if error:
                self._preview_error(error)
                return
            self.summary_var.set(
                f"工作表 {info['sheet_title']}：{info['max_row']} 行、{info['max_column']} 列；"
                f"预计生成 {info['group_count']} 个文件（包含空白关键字分组）。预览仅显示部分行列。"
            )
            preview_key = "key_row_preview" if mode == "row" else "key_column_preview"
            self.set_text_content(self.preview_text, info[preview_key])
            combo = self.key_row_combo if mode == "row" else self.key_column_combo
            combo["values"] = info[preview_key]
            if mode == "row":
                self.sync_key_row_combo()
            else:
                self.sync_key_column_combo()
            self.render_preview_table(info["column_headers"], info["preview_rows"])
            self.render_split_preview_table(info["split_objects"])

        self._start_job("preview", lambda: preview_function(Path(input_path), sheet_name, leading, key, trailing), done)

    def _preview_error(self, error):
        self.summary_var.set(f"无法预览：{error}")
        self.set_text_content(self.preview_text, [str(error)])
        self.clear_preview_table()
        self.clear_split_preview_table()

    def set_text_content(self, widget: tk.Text, lines: List[str]) -> None:
        widget.configure(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert("1.0", "\n".join(lines))
        widget.configure(state="disabled")

    def clear_preview_table(self) -> None:
        self.preview_table.delete(*self.preview_table.get_children())
        self.preview_table["columns"] = ()

    def render_preview_table(self, column_headers: List[str], preview_rows: List[Dict[str, object]]) -> None:
        columns = ["row_no"] + [f"col_{index}" for index in range(len(column_headers))]
        self.preview_table.delete(*self.preview_table.get_children())
        self.preview_table["columns"] = columns
        self.preview_column_map = {}
        self.preview_table.heading("row_no", text="行号")
        self.preview_table.column("row_no", width=60, anchor="center", stretch=False)

        for index, header in enumerate(column_headers):
            column_id = f"col_{index}"
            self.preview_table.heading(column_id, text=header)
            self.preview_table.column(column_id, width=150, anchor="w", stretch=False)
            header_text = header.split(": ", 1)[0]
            if "列/" in header_text:
                try:
                    column_no = int(header_text.rsplit("/", 1)[1])
                    self.preview_column_map[f"#{index + 2}"] = column_no
                except ValueError:
                    pass

        for row in preview_rows:
            row_no = row["row_no"]
            values = [str(row_no)] + [str(item) for item in row["values"]]
            tags = ()
            if row["kind"] == "header":
                tags = ("header",)
            elif row["kind"] == "footer":
                tags = ("footer",)
            elif row["kind"] == "keyrow":
                tags = ("keyrow",)
            elif row["kind"] == "split":
                tags = ("split",)
            self.preview_table.insert("", "end", values=values, tags=tags)

    def clear_split_preview_table(self) -> None:
        self.split_preview_table.delete(*self.split_preview_table.get_children())

    def sync_key_column_combo(self) -> None:
        current = self.key_column_var.get().strip()
        values = list(self.key_column_combo.cget("values"))
        if not current or not values:
            return
        for value in values:
            if value.startswith(f"{current}. "):
                self.key_column_combo.set(value)
                return

    def sync_key_row_combo(self) -> None:
        current = self.key_row_var.get().strip()
        values = list(self.key_row_combo.cget("values"))
        if not current or not values:
            return
        for value in values:
            if value.startswith(f"{current}. "):
                self.key_row_combo.set(value)
                return

    def render_split_preview_table(self, split_objects: List[Dict[str, object]]) -> None:
        self.clear_split_preview_table()
        if not split_objects:
            self.split_preview_table.insert("", "end", values=("当前条件下没有识别到对象", "-"))
            return
        for item in split_objects:
            self.split_preview_table.insert("", "end", values=(item["name"], item["count"]))

    def parse_int(self, value: str, fallback: Optional[int] = None) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            if fallback is not None:
                return fallback
            raise

    def _set_split_busy(self, busy):
        self._splitting = busy
        if busy:
            self._disabled_widgets = []
            def visit(widget):
                for child in widget.winfo_children():
                    if isinstance(child, (ttk.Entry, ttk.Combobox, ttk.Button, ttk.Radiobutton)):
                        self._disabled_widgets.append((child, child.cget("state")))
                        child.configure(state="disabled")
                    visit(child)
            visit(self.root)
        else:
            for widget, state in self._disabled_widgets:
                widget.configure(state=state)

    def run_split(self) -> None:
        if self._splitting:
            return
        input_text = self.input_var.get().strip()
        if not input_text:
            messagebox.showwarning("缺少文件", "请先选择待拆分的 Excel 文件。")
            return
        mode = self.mode_var.get()
        try:
            if mode == "row":
                leading, trailing, key = map(int, (self.header_cols_var.get(), self.footer_cols_var.get(), self.key_row_var.get()))
                split_function = split_workbook_by_row
            else:
                leading, trailing, key = map(int, (self.header_rows_var.get(), self.footer_rows_var.get(), self.key_column_var.get()))
                split_function = split_workbook
        except ValueError:
            messagebox.showwarning("参数错误", "固定行列数和关键行列序号必须是整数。")
            return
        input_path = Path(input_text)
        sheet = self.sheet_var.get().strip() or None
        output_text = self.output_var.get().strip()
        output_dir = Path(output_text) if output_text else None
        self._versions["preview"] = self._versions.get("preview", 0) + 1
        self._versions["load"] = self._versions.get("load", 0) + 1
        self._set_split_busy(True)
        self.summary_var.set("正在拆分并保存文件，请稍候……")

        def done(files, error):
            self._set_split_busy(False)
            if error:
                self.summary_var.set(f"拆分失败：{error}")
                messagebox.showerror("拆分失败", str(error))
                return
            output_parent = files[0].parent
            self.summary_var.set(f"拆分完成，共生成 {len(files)} 个文件，输出目录：{output_parent}")
            self.clear_split_preview_table()
            for path in files:
                self.split_preview_table.insert("", "end", values=(path.name, "已生成"))
            summary = f"已生成 {len(files)} 个文件。\n输出目录：{output_parent}\n公式将在 Excel/WPS 打开时重算。"
            if getattr(files, "warnings", None):
                self.summary_var.set(f"已生成 {len(files)} 个文件，其中 {len(files.warnings)} 个文件含引用错误，请核对。")
                messagebox.showwarning("拆分完成，需检查公式", summary + "\n\n" + "\n".join(files.warnings[:10]))
            else:
                messagebox.showinfo("拆分完成", summary)

        self._start_job("split", lambda: split_function(input_path, sheet, leading, key, output_dir, trailing), done)

    def run(self) -> None:
        self.root.mainloop()


def run_cli(args: argparse.Namespace) -> None:
    if args.mode == "row":
        files = split_workbook_by_row(
            input_path=Path(args.input),
            sheet_name=args.sheet,
            header_cols=args.header_cols,
            key_row=args.key_row,
            output_dir=Path(args.output_dir) if args.output_dir else None,
            footer_cols=getattr(args, "footer_cols", 0) or 0,
        )
    else:
        files = split_workbook(
            input_path=Path(args.input),
            sheet_name=args.sheet,
            header_rows=args.header_rows,
            key_column=args.key_column,
            output_dir=Path(args.output_dir) if args.output_dir else None,
            footer_rows=getattr(args, "footer_rows", 0) or 0,
        )
    summary = "\n".join(str(path) for path in files)
    print(f"拆分完成，共生成 {len(files)} 个文件：\n{summary}")
    for warning in getattr(files, "warnings", []):
        print(f"公式提示：{warning}", file=sys.stderr)


def main() -> None:
    args = parse_args()
    if args.input:
        try:
            run_cli(args)
        except Exception as exc:
            print(f"拆分失败：{exc}", file=sys.stderr)
            raise SystemExit(1)
        return
    SplitterApp().run()


if __name__ == "__main__":
    main()
